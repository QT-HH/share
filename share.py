import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers
from datetime import datetime, timedelta

# 전 달의 마지막 날 계산
today = datetime.today()
first_day_of_this_month = today.replace(day=1)
last_day_of_last_month = first_day_of_this_month - timedelta(days=1)
end_date = last_day_of_last_month.strftime('%Y-%m-%d')

tickerList = ["QQQ", "SOXX", "SPY", "DIA"]

# 데이터 다운로드 및 처리
def createPivotXLSX(ticker, writer):
    start_date = "2004-01-01"
    data = yf.download(ticker, start=start_date, end=end_date)
    monthly_data = data['Adj Close'].resample('ME').last()
    print(data)
    print(monthly_data)
    monthly_returns = monthly_data.pct_change() * 100  # 월간 변동률 계산

    # 피벗 테이블 생성
    monthly_returns_df = monthly_returns.reset_index()
    monthly_returns_df['Year'] = monthly_returns_df['Date'].dt.year
    monthly_returns_df['Month'] = monthly_returns_df['Date'].dt.month
    pivot_df = monthly_returns_df.pivot(index='Year', columns='Month', values='Adj Close')
    pivot_df.columns = [f"{month}월" for month in pivot_df.columns]

    # 연간 변동률 계산 (각 월간 변동률을 모두 곱한 후 최종적으로 소수점 두 번째 자리로 포맷팅)
    def calculate_yearly_return(row):
        cumulative_return = 1.0
        for value in row.dropna():
            cumulative_return *= (1 + value / 100)
        final_return = (cumulative_return - 1) * 100
        return round(final_return, 2)

    pivot_df['연간(%)'] = pivot_df.apply(calculate_yearly_return, axis=1)

    # 피벗 데이터를 지정된 시트에 저장
    pivot_df.to_excel(writer, sheet_name=ticker)

# 엑셀 파일 생성 및 데이터 작성
excel_file_path = "share.xlsx"
with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
    for ticker in tickerList:
        createPivotXLSX(ticker, writer)

# 엑셀 파일 불러오기 및 서식 적용
wb = load_workbook(excel_file_path)

for ticker in tickerList:
    ws = wb[ticker]

    gray_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for row in ws['A2:A{}'.format(ws.max_row)]:
        for cell in row:
            cell.fill = gray_fill

    for col in ws['B1:{}1'.format(chr(64 + ws.max_column))]:
        for cell in col:
            cell.fill = gray_fill

    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    red_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = '0.00'
                if cell.value < 0:
                    cell.fill = blue_fill
                elif cell.value > 0:
                    cell.fill = red_fill

# 엑셀 파일 저장
wb.save(excel_file_path)
wb.close()

print(f"포맷이 적용된 엑셀 파일이 저장되었습니다: {excel_file_path}")
